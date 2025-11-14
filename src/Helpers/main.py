import pandas as pd
from io import StringIO
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import numbers
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import Alignment
from src.Helpers import consolidation
import os
import re
from src.UI import init_logging
from src.helpers import empty_directory


developer_logger, user_logger = init_logging()

cwd = os.getcwd()
project_root = os.path.dirname(cwd)

report_dir_path1 = os.path.abspath(os.path.join(cwd, 'Output_File', 'Report_Files'))
report_dir_path = os.path.abspath(os.path.join(cwd, 'Output_File', 'Data_Files'))
print("Report Path:", report_dir_path)
print("Data Path:", report_dir_path1)



developer_logger.info(f"Tower Name: PSP")
user_logger.info(f"Tower Name: PSP")
developer_logger.info(f"Sub Function: Capital Projects")
user_logger.info(f"Sub Function: Capital Projects")
developer_logger.info(f"Use Case: Automation of invoice pending approval")
user_logger.info(f"Use Case: Automation of invoice pending approval")


def cap_pro():


    global master, mail

    cwd = os.getcwd()
    project_root = os.path.dirname(cwd)
    input_dir_path = os.path.join(project_root, 'input_folders', 'Purchase_Order_Files')

    input_dir2 = os.path.join(project_root, 'validation_input')

# ---------------purchase order file merged

    # input_dir_path = r'C:\Users\pjvg\OneDrive - WBA\PycharmProjects\input_files\Capital Projects\Invoice Report\Invoice Report\Raw Data\Purchase Order Files'
    master = next((os.path.join(input_dir2, f) for f in os.listdir(input_dir2) if f.endswith('.xlsx')), None)

    today_str = datetime.today().strftime('%m.%d.%Y')
    empty_directory(report_dir_path)
    os.makedirs(report_dir_path, exist_ok=True)
    purchase_orderfile = f'Purchase Order {today_str}.xlsx'
    output_file = os.path.join(report_dir_path, purchase_orderfile)

    # Initialize an empty list to collect all data tables
    all_data_tables = []

    # Iterate over all .xls files in the input_folders directory
    for filename in os.listdir(input_dir_path):
        if filename.endswith('.xls'):
            file_path = os.path.join(input_dir_path, filename)

            # Read the file content
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                html_content = f.read()

            # Parse all tables
            tables = pd.read_html(StringIO(html_content))

            # Ensure there are at least 4 tables to extract headers and data
            if len(tables) >= 4:
                header_table = tables[2]
                headers = header_table.columns.tolist()

                data_table = tables[3]
                data_table.columns = headers  # Apply headers from Table 3

                # Append the data table to the list
                all_data_tables.append(data_table)

    # Concatenate all data tables into a single DataFrame
    combined_data = pd.concat(all_data_tables, ignore_index=True)

    # Save the combined data to an Excel file
    combined_data.to_excel(output_file, index=False)

    print(f"Combined data from all .xls files saved to: {output_file}")
    developer_logger.info(f"Combined data from all .xls files saved to: {output_file}")
    user_logger.info(f"Combined data from all .xls files saved to: {output_file}")
# --------------------output work
#     input_dir_path2 = r'C:\Users\pjvg\OneDrive - WBA\PycharmProjects\input_files\Capital Projects\Invoice Report\Invoice Report\Raw Data\Invoice Files'
    input_dir_path2 = os.path.join(project_root, 'input_folders','Invoice_Files')

    empty_directory(report_dir_path1)
    os.makedirs(report_dir_path1, exist_ok=True)

    today_str1 = datetime.today().strftime('%m/%d/%Y')
    output_filename = f'Ariba Invoice {today_str}.xlsx'
    output_file1 = os.path.join(report_dir_path1, output_filename)



    # Status categories
    status_sheets = {
        'Submitted': [],
        'Denied': [],
        'Rejected': [],
        'Reconciling': []
    }

    # Process each .xls file
    for filename in os.listdir(input_dir_path2):
        if filename.endswith('.xls'):
            file_path = os.path.join(input_dir_path2, filename)
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                html_content = f.read()

            tables = pd.read_html(StringIO(html_content))
            if len(tables) >= 4:
                headers = tables[2].columns.tolist()
                data_table = tables[3]
                data_table.columns = headers

                for status in status_sheets:
                    filtered = data_table[data_table['Status'] == status]
                    if not filtered.empty:
                        status_sheets[status].append(filtered)

    with pd.ExcelWriter(output_file1, engine='openpyxl') as writer:
        if status_sheets['Submitted']:
            pd.concat(status_sheets['Submitted'], ignore_index=True).to_excel(writer, sheet_name='Invoice Pending Approval', index=False)
        if status_sheets['Denied']:
            pd.concat(status_sheets['Denied'], ignore_index=True).to_excel(writer, sheet_name='Denied Invoice', index=False)
        if status_sheets['Rejected']:
            pd.concat(status_sheets['Rejected'], ignore_index=True).to_excel(writer, sheet_name='Rejected Invoice', index=False)
        if status_sheets['Reconciling']:
            pd.concat(status_sheets['Reconciling'], ignore_index=True).to_excel(writer, sheet_name='Reconciling Invoice', index=False)

    print(f"Invoice report saved to: {output_file1}")
    developer_logger.info(f"Invoice report saved to: {output_file1}")
    user_logger.info(f"Invoice report saved to: {output_file1}")
#------------------------'Invoice Pending Approval'

    # Load the workbook
    wb = load_workbook(output_file1)
    sheet_name = 'Invoice Pending Approval'

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Set header row height to 30.00
        ws.row_dimensions[1].height = 30.0

        # Apply yellow fill to headers in columns A to J
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for col in range(1, 11):
            ws.cell(row=1, column=col).fill = yellow_fill

        # Format date columns to short date format
        date_columns = ['Invoice Date', 'Date Created', 'Date Submitted']
        header_row = [cell.value for cell in ws[1]]

        for col_idx, header in enumerate(header_row, start=1):
            if header in date_columns:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    # Convert string to datetime if needed
                    if isinstance(cell.value, str):
                        try:
                            cell.value = datetime.strptime(cell.value, '%m/%d/%Y %I:%M %p')
                        except ValueError:
                            continue  # Skip if not a valid datetime string
                    # Apply short date format
                    if isinstance(cell.value, datetime):
                        cell.number_format = numbers.FORMAT_DATE_XLSX14

        # Clean and format 'Invoice Amount' column
        if 'Invoice Amount' in header_row:
            amount_col_idx = header_row.index('Invoice Amount') + 1
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=amount_col_idx)
                if isinstance(cell.value, str):
                    # Remove 'USD' and any non-numeric characters except dot and comma
                    cleaned_value = re.sub(r'[^0-9.,-]', '', cell.value)
                    try:
                        numeric_value = float(cleaned_value.replace(',', ''))
                        cell.value = numeric_value
                        cell.number_format = '_($* #,##0.00_)'
                    except ValueError:
                        continue


        # Define new headers starting from column K (11)
        new_headers = [
            today_str1, 'Days Pending', 'Payment Terms', 'EMO', 'Term Days',
            'Payment Due Date', 'Preparer', 'PO Title', 'PO Status',
            'PO Created', 'FY PO Created', 'FY Program'
        ]

        bold_font = Font(bold=True)


        # Apply light green fill to new headers
        light_green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        highlight_green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        for i, header in enumerate(new_headers, start=11):
            cell = ws.cell(row=1, column=i, value=header)
            cell.fill = highlight_green_fill
            cell.font = bold_font
        # Save the workbook
        wb.save(output_file1)

    print(f"Updated '{sheet_name}' sheet with new columns.")
    developer_logger.info(f"Updated '{sheet_name}' sheet with new columns.")
    user_logger.info(f"Updated '{sheet_name}' sheet with new columns.")
# ------------------------col-L
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Find the column index for 'Days Pending'
        header_row = 1
        days_pending_col = None
        for cell in ws[header_row]:
            if cell.value == 'Days Pending':
                days_pending_col = cell.column
                break

        # Apply formula if column found
        if days_pending_col:
            for row in range(2, ws.max_row + 1):
                formula = f'=NETWORKDAYS(F{row}, $K$1)'
                ws.cell(row=row, column=days_pending_col, value=formula)

            wb.save(output_file1)
            print(f"Formula applied to 'Days Pending' column in '{sheet_name}' sheet.")
            developer_logger.info(f"Formula applied to 'Days Pending' column in '{sheet_name}' sheet.")
            user_logger.info(f"Formula applied to 'Days Pending' column in '{sheet_name}' sheet.")
        else:
            print("Column 'Days Pending' not found.")
            developer_logger.info("Column 'Days Pending' not found.")
            user_logger.info("Column 'Days Pending' not found.")
    else:
        print(f"Sheet '{sheet_name}' not found.")
        developer_logger.info(f"Sheet '{sheet_name}' not found.")
        user_logger.info(f"Sheet '{sheet_name}' not found.")
# -------------col-M

    ws = wb[sheet_name]
    master_wb = load_workbook(master)
    master_ws = master_wb.active

    # Find columns in master sheet
    start_col = end_col = None
    for cell in master_ws[1]:
        if cell.value == 'Name':
            start_col = cell.column
        if cell.value == 'Payment term text':
            end_col = cell.column
        if start_col and end_col:
            break

    start_col_letter = master_ws.cell(row=1, column=start_col).column_letter
    end_col_letter = master_ws.cell(row=1, column=end_col).column_letter
    reference_range = f"'{os.path.basename(master)}'!${start_col_letter}:${end_col_letter}"

    col_index_num = end_col - start_col + 1

    # Find columns in invoice sheet
    supplier_col = payment_terms_col = None
    for cell in ws[1]:
        if cell.value == 'Supplier':
            supplier_col = cell.column
        if cell.value == 'Payment Terms':
            payment_terms_col = cell.column

    # Apply VLOOKUP formula
    if supplier_col and payment_terms_col:
        for row in range(2, ws.max_row + 1):
            lookup_cell = ws.cell(row=row, column=supplier_col).coordinate
            formula = f'=VLOOKUP({lookup_cell}, {reference_range}, {col_index_num}, 0)'
            ws.cell(row=row, column=payment_terms_col, value=formula)

    print("VLOOKUP applied to 'Payment Terms' column.")
    developer_logger.info("VLOOKUP applied to 'Payment Terms' column.")
    user_logger.info("VLOOKUP applied to 'Payment Terms' column.")
    # ---------#N\A

    # Find columns in master sheet
    name_col = payment_term_col = None
    for cell in master_ws[1]:
        if cell.value == 'Name':
            name_col = cell.column
        if cell.value == 'Payment term text':
            payment_term_col = cell.column
        if name_col and payment_term_col:
            break

    vendor_terms = {}
    for row in range(2, master_ws.max_row + 1):
        name = master_ws.cell(row=row, column=name_col).value
        term = master_ws.cell(row=row, column=payment_term_col).value
        if name:
            parts = str(name).strip().split()
            if len(parts) >= 2:
                key = (parts[0].lower(), parts[1][:3].lower())
            elif len(parts) == 1:
                key = (parts[0].lower(), '')
            else:
                continue
            vendor_terms[key] = term

    if supplier_col and payment_terms_col:
        for row in range(2, ws.max_row + 1):
            supplier_name = ws.cell(row=row, column=supplier_col).value
            payment_cell = ws.cell(row=row, column=payment_terms_col)

            if supplier_name:
                parts = str(supplier_name).strip().split()
                if len(parts) >= 2:
                    supplier_key = (parts[0].lower(), parts[1][:3].lower())
                elif len(parts) == 1:
                    supplier_key = (parts[0].lower(), '')
                else:
                    continue

                matched_term = vendor_terms.get(supplier_key)

                if matched_term:
                    payment_cell.value = matched_term
                else:
                    payment_cell.value = "Not Found"
    # Save workbook
    wb.save(output_file1)
    print("✅ Payment Terms updated using first two-word match from master file.")
    developer_logger.info("✅ Payment Terms updated using first two-word match from master file.")
    user_logger.info("✅ Payment Terms updated using first two-word match from master file.")
# ------------------------col-N
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Find the column index for 'EMO'
        header_row = 1
        emo_col = None
        for cell in ws[header_row]:
            if cell.value == 'EMO':
                emo_col = cell.column
                break

        # Apply formula if column found
        if emo_col:
            for row in range(2, ws.max_row + 1):
                formula = f'=ISNUMBER(SEARCH(" EOM ", " " & M{row} & " "))'
                ws.cell(row=row, column=emo_col, value=formula)

            wb.save(output_file1)
            print(f"✅ Formula applied to 'EMO' column in '{sheet_name}' sheet.")
            developer_logger.info(f"✅ Formula applied to 'EMO' column in '{sheet_name}' sheet.")
            user_logger.info(f"✅ Formula applied to 'EMO' column in '{sheet_name}' sheet.")
        else:
            print("❌ Column 'EMO' not found.")
            developer_logger.info("❌ Column 'EMO' not found.")
            user_logger.info("❌ Column 'EMO' not found.")
    else:
        print(f"❌ Sheet '{sheet_name}' not found.")
        developer_logger.info(f"❌ Sheet '{sheet_name}' not found.")
        user_logger.info(f"❌ Sheet '{sheet_name}' not found.")
# ------------------------col-0

    def Term_Days():



        ws = wb[sheet_name]

        # Find column indices for 'Payment Terms' and 'Term Days'
        payment_terms_col = term_days_col = None
        for cell in ws[1]:
            if cell.value == 'Payment Terms':
                payment_terms_col = cell.column
            if cell.value == 'Term Days':
                term_days_col = cell.column

        # Extract numeric values from 'Payment Terms' and write to 'Term Days'
        if payment_terms_col and term_days_col:
            for row in range(2, ws.max_row + 1):
                payment_value = ws.cell(row=row, column=payment_terms_col).value
                term_days_cell = ws.cell(row=row, column=term_days_col)

                if payment_value:
                    numbers = re.findall(r'\d+', str(payment_value))
                    if numbers:
                        term_days_cell.value = int(numbers[-1])  # Use the last number found
                    else:
                        term_days_cell.value = None  # No number found
        else:
            print("❌ Required columns 'Payment Terms' or 'Term Days' not found.")
            developer_logger.info("❌ Required columns 'Payment Terms' or 'Term Days' not found.")
            user_logger.info("❌ Required columns 'Payment Terms' or 'Term Days' not found.")

        # Save the workbook
        wb.save(output_file1)  # Replace with desired output file name
        print("✅ 'Term Days' column updated.")
        developer_logger.info("✅ 'Term Days' column updated.")
        user_logger.info("✅ 'Term Days' column updated.")
    Term_Days()

 # ------------------------col-P

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Find the column index for 'payment due date'
        header_row = 1
        paydd_col = None
        for cell in ws[header_row]:
            if cell.value == 'Payment Due Date':
                paydd_col = cell.column
                break

        # Apply formula if column found
        if paydd_col:
            for row in range(2, ws.max_row + 1):
                formula = f'=IF(N{row}=TRUE,(EOMONTH(B{row},0))+O{row},B{row}+O{row})'
                ws.cell(row=row, column=paydd_col, value=formula)

                # Format date columns to short date format
                date_columns = ['Payment Due Date']
                header_row = [cell.value for cell in ws[1]]
                for col_idx, header in enumerate(header_row, start=1):
                    if header in date_columns:
                        for row in range(2, ws.max_row + 1):
                            ws.cell(row=row, column=col_idx).number_format = numbers.FORMAT_DATE_XLSX14
            wb.save(output_file1)
            print(f"✅ Formula applied to 'payment due date' column in '{sheet_name}' sheet.")
            developer_logger.info(f"✅ Formula applied to 'payment due date' column in '{sheet_name}' sheet.")
            user_logger.info(f"✅ Formula applied to 'payment due date' column in '{sheet_name}' sheet.")
        else:
            print("❌ Column 'payment due date' not found.")
            developer_logger.info("❌ Column 'payment due date' not found.")
            user_logger.info("❌ Column 'payment due date' not found.")
    else:
        print(f"❌ Sheet '{sheet_name}' not found.")
        developer_logger.info(f"❌ Sheet '{sheet_name}' not found.")
        user_logger.info(f"❌ Sheet '{sheet_name}' not found.")


# ------------------------col-Q


# Load purchase order workbook


    report_dir_path2 = os.path.join(cwd, 'Output_File', 'Data_Files')
    purchase_order_path = os.path.join(report_dir_path2, purchase_orderfile)
    purchase_wb = load_workbook(purchase_order_path)
    purchase_ws = purchase_wb.active

    # Identify 'Order ID' column
    header = [cell.value for cell in purchase_ws[1]]
    order_id_col = header.index('Order ID') + 1

    # Convert scientific notation to string format
    for row in range(2, purchase_ws.max_row + 1):
        cell = purchase_ws.cell(row=row, column=order_id_col)
        value = cell.value
        if value is not None:
            try:
                # Convert float/scientific to int then to string
                normalized = str(int(value))
                cell.value = normalized
                cell.number_format = '@'  # Set format to Text
            except (ValueError, TypeError):
                continue

    # Save the updated file
    # purchase_wb.save(output_file)
    print("✅ 'Order ID' column updated to prevent scientific notation display.")
    developer_logger.info("✅ 'Order ID' column updated to prevent scientific notation display.")
    user_logger.info("✅ 'Order ID' column updated to prevent scientific notation display.")


    # Convert text values back to number format
    for row in range(2, purchase_ws.max_row + 1):
        cell = purchase_ws.cell(row=row, column=order_id_col)
        value = cell.value
        if value is not None:
            try:
                numeric_value = int(value)  # Convert string to integer
                cell.value = numeric_value
                cell.number_format = '0'  # Format as number with no decimals
            except (ValueError, TypeError):
                continue

    # Save the updated workbook
    # wb.save(purchase_order_path)

    print("✅ 'Order ID' column converted back to number format.")
    developer_logger.info("✅ 'Order ID' column converted back to number format.")
    user_logger.info("✅ 'Order ID' column converted back to number format.")
    # ----------wrap
    # Apply auto-filter
    max_col = purchase_ws.max_column
    last_col_letter = purchase_ws.cell(row=1, column=max_col).column_letter
    purchase_ws.auto_filter.ref = f"A1:{last_col_letter}1"

    # Wrap text in header cells
    for col in range(1, max_col + 1):
        cell = purchase_ws.cell(row=1, column=col)
        cell.alignment = Alignment(wrap_text=True)

    # Format date columns to short date format
    date_column = ['Date Created']
    head_row = [cell.value for cell in purchase_ws[1]]

    for col_idx, header in enumerate(head_row, start=1):
        if header in date_column:
            for row in range(2, purchase_ws.max_row + 1):
                cell = purchase_ws.cell(row=row, column=col_idx)
                # Convert string to datetime if needed
                if isinstance(cell.value, str):
                    try:
                        cell.value = datetime.strptime(cell.value, '%m/%d/%Y %I:%M %p')
                    except ValueError:
                        continue  # Skip if not a valid datetime string
                # Apply short date format
                if isinstance(cell.value, datetime):
                    cell.number_format = numbers.FORMAT_DATE_XLSX14

    # Adjust column widths based on content
    for col in purchase_ws.iter_cols(min_row=1, max_row=purchase_ws.max_row, max_col=max_col):
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                max_length = max(max_length, length)

        # Non-linear approximation: base + scaled length
        if max_length <= 5:
            width = 7
        elif max_length <= 10:
            width = 14
        elif max_length <= 20:
            width = 22
        elif max_length <= 30:
            width = 30
        else:
            width = min(50, max_length * 0.9)  # Cap at 25

        purchase_ws.column_dimensions[column].width = width

    purchase_wb.save(output_file)
    print(f"Auto-filters, wrapped headers, and column widths adjusted in '{sheet_name}' sheet.")
    developer_logger.info(f"Auto-filters, wrapped headers, and column widths adjusted in '{sheet_name}' sheet.")
    user_logger.info(f"Auto-filters, wrapped headers, and column widths adjusted in '{sheet_name}' sheet.")





    def preparer():
        # Find columns in purchase order sheet
        start_col1 = end_col1 = None
        for cell in purchase_ws[1]:
            if cell.value == 'Order ID':
                start_col1 = cell.column
            elif cell.value == 'Preparer':
                end_col1 = cell.column
            if start_col1 and end_col1:
                break

        start_col_letter1 = purchase_ws.cell(row=1, column=start_col1).column_letter
        end_col_letter1 = purchase_ws.cell(row=1, column=end_col1).column_letter
        reference_range1 = f"'{os.path.basename(purchase_order_path)}'!${start_col_letter1}:${end_col_letter1}"
        col_index_num1 = end_col1 - start_col1 + 1

        # Find columns in invoice sheet
        Ini_Match_col = Preparer_col = Supplier_col = None
        for cell in ws[1]:
            if cell.value == 'Initially Matched Order':
                Ini_Match_col = cell.column
            elif cell.value == 'Preparer':
                Preparer_col = cell.column
            elif cell.value == 'Supplier':
                Supplier_col = cell.column

        # Apply VLOOKUP formula to Preparer column
        if Ini_Match_col and Preparer_col:
            for row in range(2, ws.max_row + 1):
                lookup_cell = ws.cell(row=row, column=Ini_Match_col).coordinate
                formula = f'=VLOOKUP({lookup_cell}, {reference_range1}, {col_index_num1}, 0)'
                ws.cell(row=row, column=Preparer_col, value=formula)

        print("✅ VLOOKUP applied to 'Preparer' column.")
        developer_logger.info("✅ VLOOKUP applied to 'Preparer' column.")
        user_logger.info("✅ VLOOKUP applied to 'Preparer' column.")

        # Build fallback dictionary from purchase order sheet
        name_col = preparer_col_master = None
        for cell in purchase_ws[1]:
            if cell.value == 'Supplier Name':
                name_col = cell.column
            elif cell.value == 'Preparer':
                preparer_col_master = cell.column
            if name_col and preparer_col_master:
                break

        vendor_preparer = {}
        for row in range(2, purchase_ws.max_row + 1):
            name = purchase_ws.cell(row=row, column=name_col).value
            preparer = purchase_ws.cell(row=row, column=preparer_col_master).value
            if name:
                key = ' '.join(str(name).strip().split()[:2]).lower()
                vendor_preparer[key] = preparer

        # Apply fallback if VLOOKUP fails
        if Supplier_col and Preparer_col:
            for row in range(2, ws.max_row + 1):
                preparer_cell = ws.cell(row=row, column=Preparer_col)
                if preparer_cell.value == '#N/A' or preparer_cell.value is None:
                    supplier_name = ws.cell(row=row, column=Supplier_col).value
                    if supplier_name:
                        supplier_key = ' '.join(str(supplier_name).strip().split()[:2]).lower()
                        matched_preparer = vendor_preparer.get(supplier_key)
                        if matched_preparer:
                            preparer_cell.value = matched_preparer
                        else:
                            preparer_cell.value = "Not Found"

        # Save workbook
        wb.save(output_file1)
        print("✅ Preparer column updated with fallback logic.")
        developer_logger.info("✅ Preparer column updated with fallback logic.")
        user_logger.info("✅ Preparer column updated with fallback logic.")
    preparer()

# ------------------------col-R

    def PO_title():
        # Find columns in purchase order sheet
        start_col1 = end_col1 = None
        for cell in purchase_ws[1]:
            if cell.value == 'Order ID':
                start_col1 = cell.column
            elif cell.value == 'Title':
                end_col1 = cell.column
            if start_col1 and end_col1:
                break

        start_col_letter1 = purchase_ws.cell(row=1, column=start_col1).column_letter
        end_col_letter1 = purchase_ws.cell(row=1, column=end_col1).column_letter
        reference_range1 = f"'{os.path.basename(purchase_order_path)}'!${start_col_letter1}:${end_col_letter1}"
        col_index_num1 = end_col1 - start_col1 + 1

        # Find columns in invoice sheet
        Ini_Match_col = Preparer_col = Supplier_col = None
        for cell in ws[1]:
            if cell.value == 'Initially Matched Order':
                Ini_Match_col = cell.column
            elif cell.value == 'PO Title':
                Preparer_col = cell.column
            elif cell.value == 'Supplier':
                Supplier_col = cell.column

        # Apply VLOOKUP formula to Preparer column
        if Ini_Match_col and Preparer_col:
            for row in range(2, ws.max_row + 1):
                lookup_cell = ws.cell(row=row, column=Ini_Match_col).coordinate
                formula = f'=VLOOKUP({lookup_cell}, {reference_range1}, {col_index_num1}, 0)'
                ws.cell(row=row, column=Preparer_col, value=formula)

        print("✅ VLOOKUP applied to 'PO Title' column.")
        developer_logger.info("✅ VLOOKUP applied to 'PO Title' column.")
        user_logger.info("✅ VLOOKUP applied to 'PO Title' column.")

        # Build fallback dictionary from purchase order sheet
        name_col = preparer_col_master = None
        for cell in purchase_ws[1]:
            if cell.value == 'Supplier Name':
                name_col = cell.column
            elif cell.value == 'Title':
                preparer_col_master = cell.column
            if name_col and preparer_col_master:
                break

        vendor_preparer = {}
        for row in range(2, purchase_ws.max_row + 1):
            name = purchase_ws.cell(row=row, column=name_col).value
            preparer = purchase_ws.cell(row=row, column=preparer_col_master).value
            if name:
                key = ' '.join(str(name).strip().split()[:2]).lower()
                vendor_preparer[key] = preparer

        # Apply fallback if VLOOKUP fails
        if Supplier_col and Preparer_col:
            for row in range(2, ws.max_row + 1):
                preparer_cell = ws.cell(row=row, column=Preparer_col)
                if preparer_cell.value == '#N/A' or preparer_cell.value is None:
                    supplier_name = ws.cell(row=row, column=Supplier_col).value
                    if supplier_name:
                        supplier_key = ' '.join(str(supplier_name).strip().split()[:2]).lower()
                        matched_preparer = vendor_preparer.get(supplier_key)
                        if matched_preparer:
                            preparer_cell.value = matched_preparer
                        else:
                            preparer_cell.value = "Not Found"

        # Save workbook
        wb.save(output_file1)
        print("✅ PO Title column updated with fallback logic.")
        developer_logger.info("✅ PO Title column updated with fallback logic.")
        user_logger.info("✅ PO Title column updated with fallback logic.")


    PO_title()
# ------------------------col-S

    def PO_Status():
        # Find columns in purchase order sheet
        start_col1 = end_col1 = None
        for cell in purchase_ws[1]:
            if cell.value == 'Order ID':
                start_col1 = cell.column
            elif cell.value == 'Status':
                end_col1 = cell.column
            if start_col1 and end_col1:
                break

        start_col_letter1 = purchase_ws.cell(row=1, column=start_col1).column_letter
        end_col_letter1 = purchase_ws.cell(row=1, column=end_col1).column_letter
        reference_range1 = f"'{os.path.basename(purchase_order_path)}'!${start_col_letter1}:${end_col_letter1}"
        col_index_num1 = end_col1 - start_col1 + 1

        # Find columns in invoice sheet
        Ini_Match_col = Preparer_col = Supplier_col = None
        for cell in ws[1]:
            if cell.value == 'Initially Matched Order':
                Ini_Match_col = cell.column
            elif cell.value == 'PO Status':
                Preparer_col = cell.column
            elif cell.value == 'Supplier':
                Supplier_col = cell.column

        # Apply VLOOKUP formula to Preparer column
        if Ini_Match_col and Preparer_col:
            for row in range(2, ws.max_row + 1):
                lookup_cell = ws.cell(row=row, column=Ini_Match_col).coordinate
                formula = f'=VLOOKUP({lookup_cell}, {reference_range1}, {col_index_num1}, 0)'
                ws.cell(row=row, column=Preparer_col, value=formula)

        print("✅ VLOOKUP applied to 'PO Status' column.")
        developer_logger.info("✅ VLOOKUP applied to 'PO Status' column.")
        user_logger.info("✅ VLOOKUP applied to 'PO Status' column.")

        # Build fallback dictionary from purchase order sheet
        name_col = preparer_col_master = None
        for cell in purchase_ws[1]:
            if cell.value == 'Supplier Name':
                name_col = cell.column
            elif cell.value == 'Status':
                preparer_col_master = cell.column
            if name_col and preparer_col_master:
                break

        vendor_preparer = {}
        for row in range(2, purchase_ws.max_row + 1):
            name = purchase_ws.cell(row=row, column=name_col).value
            preparer = purchase_ws.cell(row=row, column=preparer_col_master).value
            if name:
                key = ' '.join(str(name).strip().split()[:2]).lower()
                vendor_preparer[key] = preparer

        # Apply fallback if VLOOKUP fails
        if Supplier_col and Preparer_col:
            for row in range(2, ws.max_row + 1):
                preparer_cell = ws.cell(row=row, column=Preparer_col)
                if preparer_cell.value == '#N/A' or preparer_cell.value is None:
                    supplier_name = ws.cell(row=row, column=Supplier_col).value
                    if supplier_name:
                        supplier_key = ' '.join(str(supplier_name).strip().split()[:2]).lower()
                        matched_preparer = vendor_preparer.get(supplier_key)
                        if matched_preparer:
                            preparer_cell.value = matched_preparer
                        else:
                            preparer_cell.value = "Not Found"

        # Save workbook
        wb.save(output_file1)
        print("✅ PO Status column updated with fallback logic.")
        developer_logger.info("✅ PO Status column updated with fallback logic.")
        user_logger.info("✅ PO Status column updated with fallback logic.")


    PO_Status()
# ------------------------col-T

    def PO_Created():
        # Find columns in purchase order sheet
        start_col1 = end_col1 = None
        for cell in purchase_ws[1]:
            if cell.value == 'Order ID':
                start_col1 = cell.column
            elif cell.value == 'Date Created':
                end_col1 = cell.column
            if start_col1 and end_col1:
                break

        start_col_letter1 = purchase_ws.cell(row=1, column=start_col1).column_letter
        end_col_letter1 = purchase_ws.cell(row=1, column=end_col1).column_letter
        reference_range1 = f"'{os.path.basename(purchase_order_path)}'!${start_col_letter1}:${end_col_letter1}"
        col_index_num1 = end_col1 - start_col1 + 1

        # Find columns in invoice sheet
        Ini_Match_col = Preparer_col = Supplier_col = None
        for cell in ws[1]:
            if cell.value == 'Initially Matched Order':
                Ini_Match_col = cell.column
            elif cell.value == 'PO Created':
                Preparer_col = cell.column
            elif cell.value == 'Supplier':
                Supplier_col = cell.column

        # Apply VLOOKUP formula to Preparer column
        if Ini_Match_col and Preparer_col:
            for row in range(2, ws.max_row + 1):
                lookup_cell = ws.cell(row=row, column=Ini_Match_col).coordinate
                formula = f'=VLOOKUP({lookup_cell}, {reference_range1}, {col_index_num1}, 0)'
                ws.cell(row=row, column=Preparer_col, value=formula)

        print("✅ VLOOKUP applied to 'PO Created' column.")
        developer_logger.info("✅ VLOOKUP applied to 'PO Created' column.")
        user_logger.info("✅ VLOOKUP applied to 'PO Created' column.")

        # Build fallback dictionary from purchase order sheet
        name_col = preparer_col_master = None
        for cell in purchase_ws[1]:
            if cell.value == 'Supplier Name':
                name_col = cell.column
            elif cell.value == 'Date Created':
                preparer_col_master = cell.column
            if name_col and preparer_col_master:
                break

        vendor_preparer = {}
        for row in range(2, purchase_ws.max_row + 1):
            name = purchase_ws.cell(row=row, column=name_col).value
            preparer = purchase_ws.cell(row=row, column=preparer_col_master).value
            if name:
                key = ' '.join(str(name).strip().split()[:2]).lower()
                vendor_preparer[key] = preparer

        # Apply fallback if VLOOKUP fails
        if Supplier_col and Preparer_col:
            for row in range(2, ws.max_row + 1):
                preparer_cell = ws.cell(row=row, column=Preparer_col)
                if preparer_cell.value == '#N/A' or preparer_cell.value is None:
                    supplier_name = ws.cell(row=row, column=Supplier_col).value
                    if supplier_name:
                        supplier_key = ' '.join(str(supplier_name).strip().split()[:2]).lower()
                        matched_preparer = vendor_preparer.get(supplier_key)
                        if matched_preparer:
                            preparer_cell.value = matched_preparer
                        else:
                            preparer_cell.value = "Not Found"

        # if sheet_name in wb.sheetnames:
        #     ws = wb[sheet_name]

        # Get header row
        header_row = [cell.value for cell in ws[1]]

        # Find the column index for "Date Created"
        if 'Date Created' in header_row:
            col_idx = header_row.index('PO Created') + 1

            # Apply short date format to all cells in that column (excluding header)
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.number_format = 'mm/dd/yyyy'


        # Save workbook
        wb.save(output_file1)
        print("✅ PO Created column updated with fallback logic.")
        developer_logger.info("✅ PO Created column updated with fallback logic.")
        user_logger.info("✅ PO Created column updated with fallback logic.")


    PO_Created()

# ------------------------col-U(FY PO Created)


    # Create a mapping of Order ID to Date Created
    purchase_header = [cell.value for cell in purchase_ws[1]]
    order_id_idx = purchase_header.index('Order ID') + 1
    date_created_idx = purchase_header.index('Date Created') + 1

    order_date_map = {}
    for row in range(2, purchase_ws.max_row + 1):
        order_id = purchase_ws.cell(row=row, column=order_id_idx).value
        date_created = purchase_ws.cell(row=row, column=date_created_idx).value
        order_date_map[order_id] = date_created  # Use raw value as-is

    # Identify columns in original worksheet
    original_header = [cell.value for cell in ws[1]]
    matched_order_idx = original_header.index('Initially Matched Order') + 1
    if 'FY PO Created' in original_header:
        fy_col_idx = original_header.index('FY PO Created') + 1
    else:
        fy_col_idx = len(original_header) + 1
        ws.cell(row=1, column=fy_col_idx, value='FY PO Created')



    def get_fiscal_year(date_obj):
        if not isinstance(date_obj, datetime):
            try:
                # Try parsing if it's a string
                date_obj = datetime.strptime(str(date_obj), "%m/%d/%Y %I:%M %p")
            except Exception:
                return None
        year = date_obj.year
        return f"FY{(year + 1) % 100:02d}" if date_obj.month >= 9 else f"FY{year % 100:02d}"

    print("Sample mapping:")
    for k, v in list(order_date_map.items())[:5]:
        print(f"{k} -> {v} ({type(v)})")

    for row in range(2, ws.max_row + 1):
        matched_order = ws.cell(row=row, column=matched_order_idx).value
        date_created = order_date_map.get(matched_order)
        fy_value = get_fiscal_year(date_created)
        if fy_value:
            ws.cell(row=row, column=fy_col_idx, value=fy_value)


    # Save the updated file
    wb.save(output_file1)
    print("✅ Fiscal year values updated based on 'Date Created' from Purchase Order.xlsx.")
    developer_logger.info("✅ Fiscal year values updated based on 'Date Created' from Purchase Order.xlsx.")
    user_logger.info("✅ Fiscal year values updated based on 'Date Created' from Purchase Order.xlsx.")



# # ------------------------col-V(FY Program)

    # Identify column indices
    purchase_header = [cell.value for cell in purchase_ws[1]]
    order_id_idx = purchase_header.index('Order ID') + 1
    title_idx = purchase_header.index('Title') + 1

    # Create mapping of Order ID to Title
    order_title_map = {}
    for row in range(2, purchase_ws.max_row + 1):
        order_id = purchase_ws.cell(row=row, column=order_id_idx).value
        title = purchase_ws.cell(row=row, column=title_idx).value
        if order_id is not None:
            order_title_map[str(order_id)] = title

    # Identify columns
    original_header = [cell.value for cell in ws[1]]
    matched_order_idx = original_header.index('Initially Matched Order') + 1
    fy_po_created_idx = original_header.index('FY PO Created') + 1
    if 'FY Program' in original_header:
        fy_program_idx = original_header.index('FY Program') + 1
    else:
        fy_program_idx = len(original_header) + 1
        ws.cell(row=1, column=fy_program_idx, value='FY Program')

    # Apply FY Program logic
    for row in range(2, ws.max_row + 1):
        matched_order = ws.cell(row=row, column=matched_order_idx).value
        fy_po_created = ws.cell(row=row, column=fy_po_created_idx).value
        fy_program_cell = ws.cell(row=row, column=fy_program_idx)

        fy_found = None
        title = order_title_map.get(str(matched_order))
        if title:
            title_str = str(title).upper()
            match = re.search(r'FY[\s\-]?(\d{2})', title_str)
            if match:
                fy_found = f"FY{match.group(1)}"

        fy_program_cell.value = fy_found if fy_found else fy_po_created

    wb.save(output_file1)
    print("✅ 'FY Program' column updated using Title from Purchase Order.xlsx.")
    developer_logger.info("✅ 'FY Program' column updated using Title from Purchase Order.xlsx.")
    user_logger.info("✅ 'FY Program' column updated using Title from Purchase Order.xlsx.")
# -------------------------------------apply filters,wraping


    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Determine the last column with headers
        max_col = ws.max_column
        last_col_letter = ws.cell(row=1, column=max_col).column_letter

        # Apply auto-filter to the header row
        ws.auto_filter.ref = f"A1:{last_col_letter}1"

        # Save the workbook
        wb.save(output_file1)
        print(f"Auto-filters applied to all headers in '{sheet_name}' sheet.")
        developer_logger.info(f"Auto-filters applied to all headers in '{sheet_name}' sheet.")
        user_logger.info(f"Auto-filters applied to all headers in '{sheet_name}' sheet.")
    else:
        print(f"Sheet '{sheet_name}' not found in the workbook.")
        developer_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")
        user_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")
# ----------wrap
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Apply auto-filter
        max_col = ws.max_column
        last_col_letter = ws.cell(row=1, column=max_col).column_letter
        ws.auto_filter.ref = f"A1:{last_col_letter}1"

        # Wrap text in header cells
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths based on content
        for col in ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=max_col):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    length = len(str(cell.value))
                    max_length = max(max_length, length)

            # Non-linear approximation: base + scaled length
            if max_length <= 5:
                width = 5
            elif max_length <= 10:
                width = 10
            elif max_length <= 20:
                width = 15
            elif max_length <= 30:
                width = 20
            elif max_length <= 35:
                width = 40
            elif max_length <= 47:
                width = 13
            else:
                width = min(50, max_length * 0.9)  # Cap at 25

            ws.column_dimensions[column].width = width

        wb.save(output_file1)
        print(f"Auto-filters, wrapped headers, and column widths adjusted in '{sheet_name}' sheet.")
        developer_logger.info(f"Auto-filters, wrapped headers, and column widths adjusted in '{sheet_name}' sheet.")
        user_logger.info(f"Auto-filters, wrapped headers, and column widths adjusted in '{sheet_name}' sheet.")
    else:
        print(f"Sheet '{sheet_name}' not found in the workbook.")
        developer_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")
        user_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")

    consolidation.cap_pro2()
