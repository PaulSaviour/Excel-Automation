from openpyxl.styles import PatternFill, Font
from openpyxl.styles import numbers
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
import re
import os
from src.UI import init_logging
from src import sendMail

developer_logger, user_logger = init_logging()

cwd = os.getcwd()
project_root = os.path.dirname(cwd)
input_dir = os.path.join(project_root, 'src')

report_dir_path = os.path.join(cwd, 'Output_File', 'Data_Files')
report_dir_path1 = os.path.join(cwd, 'Output_File', 'Report_Files')



def cap_pro2():

    def Denied_Invoice():
    #------------------------'Denied Invoice'
        today_str = datetime.today().strftime('%m.%d.%Y')
        output_filename = f'Ariba Invoice {today_str}.xlsx'
        output_file1 = os.path.join(report_dir_path1, output_filename)


        # Load the workbook
        wb = load_workbook(output_file1)
        sheet_name = 'Denied Invoice'

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Set header row height to 30.00
            ws.row_dimensions[1].height = 30.0

            # Apply yellow fill to headers in columns A to J
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for col in range(1, 11):
                ws.cell(row=1, column=col).fill = yellow_fill

            # Format date columns to short date format
            date_columns = ['Invoice Date', 'Date Created', 'Date Submitted']
            header_row = [cell.value for cell in ws[1]]
            for col_idx, header in enumerate(header_row, start=1):
                if header in date_columns:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        # Convert string to datetime if needed
                        if isinstance(cell.value, str):
                            try:
                                cell.value = datetime.strptime(cell.value, '%m/%d/%Y %I:%M %p')
                            except ValueError:
                                continue  # Skip if not a valid datetime string
                        # Apply short date format
                        if isinstance(cell.value, datetime):
                            cell.number_format = numbers.FORMAT_DATE_XLSX14
            # Clean and format 'Invoice Amount' column
            if 'Invoice Amount' in header_row:
                amount_col_idx = header_row.index('Invoice Amount') + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=amount_col_idx)
                    if isinstance(cell.value, str):
                        # Remove 'USD' and any non-numeric characters except dot and comma
                        cleaned_value = re.sub(r'[^0-9.,-]', '', cell.value)
                        try:
                            numeric_value = float(cleaned_value.replace(',', ''))
                            cell.value = numeric_value
                            cell.number_format = '_($* #,##0.00_)'
                        except ValueError:
                            continue

            # Define new headers starting from column K (11)
            new_headers = [
                'Preparer', 'PO Title', 'PO Status'
            ]
            bold_font = Font(bold=True)

            # Apply light green fill to new headers
            light_green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            highlight_green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')

            for i, header in enumerate(new_headers, start=11):
                cell = ws.cell(row=1, column=i, value=header)
                cell.fill = highlight_green_fill
                cell.font = bold_font

            # Save the workbook
            wb.save(output_file1)

            print(f"Updated '{sheet_name}' sheet with new columns.")
            developer_logger.info(f"Updated '{sheet_name}' sheet with new columns.")
            user_logger.info(f"Updated '{sheet_name}' sheet with new columns.")

            report_dir_path2 = os.path.join(cwd, 'Output_File', 'Data_Files')
            purchase_orderfile = f'Purchase Order {today_str}.xlsx'
            purchase_order_path = os.path.join(report_dir_path, purchase_orderfile)

            # purchase_order_path = os.path.join(report_dir_path2, 'purchase_orders.xlsx')
            purchase_wb = load_workbook(purchase_order_path)
            purchase_ws = purchase_wb.active

            def preparer():
                # Find columns in purchase order sheet
                start_col1 = end_col1 = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Order ID':
                        start_col1 = cell.column
                    elif cell.value == 'Preparer':
                        end_col1 = cell.column
                    if start_col1 and end_col1:
                        break

                start_col_letter1 = purchase_ws.cell(row=1, column=start_col1).column_letter
                end_col_letter1 = purchase_ws.cell(row=1, column=end_col1).column_letter
                reference_range1 = f"'{os.path.basename(purchase_order_path)}'!${start_col_letter1}:${end_col_letter1}"
                col_index_num1 = end_col1 - start_col1 + 1

                # Find columns in invoice sheet
                Ini_Match_col = Preparer_col = Supplier_col = None
                for cell in ws[1]:
                    if cell.value == 'Initially Matched Order':
                        Ini_Match_col = cell.column
                    elif cell.value == 'Preparer':
                        Preparer_col = cell.column
                    elif cell.value == 'Supplier':
                        Supplier_col = cell.column

                # Apply VLOOKUP formula to Preparer column
                if Ini_Match_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        lookup_cell = ws.cell(row=row, column=Ini_Match_col).coordinate
                        formula = f'=VLOOKUP({lookup_cell}, {reference_range1}, {col_index_num1}, 0)'
                        ws.cell(row=row, column=Preparer_col, value=formula)

                print("✅ VLOOKUP applied to 'Preparer' column.")
                developer_logger.info("✅ VLOOKUP applied to 'Preparer' column.")
                user_logger.info("✅ VLOOKUP applied to 'Preparer' column.")

                # Build fallback dictionary from purchase order sheet
                name_col = preparer_col_master = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Supplier Name':
                        name_col = cell.column
                    elif cell.value == 'Preparer':
                        preparer_col_master = cell.column
                    if name_col and preparer_col_master:
                        break

                vendor_preparer = {}
                for row in range(2, purchase_ws.max_row + 1):
                    name = purchase_ws.cell(row=row, column=name_col).value
                    preparer = purchase_ws.cell(row=row, column=preparer_col_master).value
                    if name:
                        key = ' '.join(str(name).strip().split()[:2]).lower()
                        vendor_preparer[key] = preparer

                # Apply fallback if VLOOKUP fails
                if Supplier_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        preparer_cell = ws.cell(row=row, column=Preparer_col)
                        if preparer_cell.value == '#N/A' or preparer_cell.value is None:
                            supplier_name = ws.cell(row=row, column=Supplier_col).value
                            if supplier_name:
                                supplier_key = ' '.join(str(supplier_name).strip().split()[:2]).lower()
                                matched_preparer = vendor_preparer.get(supplier_key)
                                if matched_preparer:
                                    preparer_cell.value = matched_preparer
                                else:
                                    preparer_cell.value = "Not Found"

                # Save workbook
                wb.save(output_file1)
                print("✅ Preparer column updated with fallback logic.")
                developer_logger.info("✅ Preparer column updated with fallback logic.")
                user_logger.info("✅ Preparer column updated with fallback logic.")

            preparer()

            # ------------------------col-R

            def PO_title():
                # Find columns in purchase order sheet
                start_col1 = end_col1 = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Order ID':
                        start_col1 = cell.column
                    elif cell.value == 'Title':
                        end_col1 = cell.column
                    if start_col1 and end_col1:
                        break

                start_col_letter1 = purchase_ws.cell(row=1, column=start_col1).column_letter
                end_col_letter1 = purchase_ws.cell(row=1, column=end_col1).column_letter
                reference_range1 = f"'{os.path.basename(purchase_order_path)}'!${start_col_letter1}:${end_col_letter1}"
                col_index_num1 = end_col1 - start_col1 + 1

                # Find columns in invoice sheet
                Ini_Match_col = Preparer_col = Supplier_col = None
                for cell in ws[1]:
                    if cell.value == 'Initially Matched Order':
                        Ini_Match_col = cell.column
                    elif cell.value == 'PO Title':
                        Preparer_col = cell.column
                    elif cell.value == 'Supplier':
                        Supplier_col = cell.column

                # Apply VLOOKUP formula to Preparer column
                if Ini_Match_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        lookup_cell = ws.cell(row=row, column=Ini_Match_col).coordinate
                        formula = f'=VLOOKUP({lookup_cell}, {reference_range1}, {col_index_num1}, 0)'
                        ws.cell(row=row, column=Preparer_col, value=formula)

                print("✅ VLOOKUP applied to 'PO Title' column.")
                developer_logger.info("✅ VLOOKUP applied to 'PO Title' column.")
                user_logger.info("✅ VLOOKUP applied to 'PO Title' column.")

                # Build fallback dictionary from purchase order sheet
                name_col = preparer_col_master = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Supplier Name':
                        name_col = cell.column
                    elif cell.value == 'Title':
                        preparer_col_master = cell.column
                    if name_col and preparer_col_master:
                        break

                vendor_preparer = {}
                for row in range(2, purchase_ws.max_row + 1):
                    name = purchase_ws.cell(row=row, column=name_col).value
                    preparer = purchase_ws.cell(row=row, column=preparer_col_master).value
                    if name:
                        key = ' '.join(str(name).strip().split()[:2]).lower()
                        vendor_preparer[key] = preparer

                # Apply fallback if VLOOKUP fails
                if Supplier_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        preparer_cell = ws.cell(row=row, column=Preparer_col)
                        if preparer_cell.value == '#N/A' or preparer_cell.value is None:
                            supplier_name = ws.cell(row=row, column=Supplier_col).value
                            if supplier_name:
                                supplier_key = ' '.join(str(supplier_name).strip().split()[:2]).lower()
                                matched_preparer = vendor_preparer.get(supplier_key)
                                if matched_preparer:
                                    preparer_cell.value = matched_preparer
                                else:
                                    preparer_cell.value = "Not Found"

                # Save workbook
                wb.save(output_file1)
                print("✅ PO Title column updated with fallback logic.")
                developer_logger.info("✅ PO Title column updated with fallback logic.")
                user_logger.info("✅ PO Title column updated with fallback logic.")

            PO_title()

            # ------------------------col-S

            def PO_Status():
                # Find columns in purchase order sheet
                start_col1 = end_col1 = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Order ID':
                        start_col1 = cell.column
                    elif cell.value == 'Status':
                        end_col1 = cell.column
                    if start_col1 and end_col1:
                        break

                start_col_letter1 = purchase_ws.cell(row=1, column=start_col1).column_letter
                end_col_letter1 = purchase_ws.cell(row=1, column=end_col1).column_letter
                reference_range1 = f"'{os.path.basename(purchase_order_path)}'!${start_col_letter1}:${end_col_letter1}"
                col_index_num1 = end_col1 - start_col1 + 1

                # Find columns in invoice sheet
                Ini_Match_col = Preparer_col = Supplier_col = None
                for cell in ws[1]:
                    if cell.value == 'Initially Matched Order':
                        Ini_Match_col = cell.column
                    elif cell.value == 'PO Status':
                        Preparer_col = cell.column
                    elif cell.value == 'Supplier':
                        Supplier_col = cell.column

                # Apply VLOOKUP formula to Preparer column
                if Ini_Match_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        lookup_cell = ws.cell(row=row, column=Ini_Match_col).coordinate
                        formula = f'=VLOOKUP({lookup_cell}, {reference_range1}, {col_index_num1}, 0)'
                        ws.cell(row=row, column=Preparer_col, value=formula)

                print("✅ VLOOKUP applied to 'PO Status' column.")
                developer_logger.info("✅ VLOOKUP applied to 'PO Status' column.")
                user_logger.info("✅ VLOOKUP applied to 'PO Status' column.")

                # Build fallback dictionary from purchase order sheet
                name_col = preparer_col_master = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Supplier Name':
                        name_col = cell.column
                    elif cell.value == 'Status':
                        preparer_col_master = cell.column
                    if name_col and preparer_col_master:
                        break

                vendor_preparer = {}
                for row in range(2, purchase_ws.max_row + 1):
                    name = purchase_ws.cell(row=row, column=name_col).value
                    preparer = purchase_ws.cell(row=row, column=preparer_col_master).value
                    if name:
                        key = ' '.join(str(name).strip().split()[:2]).lower()
                        vendor_preparer[key] = preparer

                # Apply fallback if VLOOKUP fails
                if Supplier_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        preparer_cell = ws.cell(row=row, column=Preparer_col)
                        if preparer_cell.value == '#N/A' or preparer_cell.value is None:
                            supplier_name = ws.cell(row=row, column=Supplier_col).value
                            if supplier_name:
                                supplier_key = ' '.join(str(supplier_name).strip().split()[:2]).lower()
                                matched_preparer = vendor_preparer.get(supplier_key)
                                if matched_preparer:
                                    preparer_cell.value = matched_preparer
                                else:
                                    preparer_cell.value = "Not Found"

                # Save workbook
                wb.save(output_file1)
                print("✅ PO Status column updated with fallback logic.")
                developer_logger.info("✅ PO Status column updated with fallback logic.")
                user_logger.info("✅ PO Status column updated with fallback logic.")

            PO_Status()
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Determine the last column with headers
                max_col = ws.max_column
                last_col_letter = ws.cell(row=1, column=max_col).column_letter

                # Apply auto-filter to the header row
                ws.auto_filter.ref = f"A1:{last_col_letter}1"

                # Save the workbook
                wb.save(output_file1)
                print(f"Auto-filters applied to all headers in '{sheet_name}' sheet.")
                developer_logger.info(f"Auto-filters applied to all headers in '{sheet_name}' sheet.")
                user_logger.info(f"Auto-filters applied to all headers in '{sheet_name}' sheet.")
            else:
                print(f"Sheet '{sheet_name}' not found in the workbook.")
                developer_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")
                user_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")
        else:
            print(f"Sheet '{sheet_name}' not found in {output_filename}.")
            developer_logger.info(f"Sheet '{sheet_name}' not found in {output_filename}.")
            user_logger.info(f"Sheet '{sheet_name}' not found in {output_filename}.")


        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Apply auto-filter
            max_col = ws.max_column
            last_col_letter = ws.cell(row=1, column=max_col).column_letter
            ws.auto_filter.ref = f"A1:{last_col_letter}1"

            # Wrap text in header cells
            for col in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col)
                cell.alignment = Alignment(wrap_text=True)

            # Adjust column widths based on content
            for col in ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=max_col):
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        length = len(str(cell.value))
                        max_length = max(max_length, length)

                # Non-linear approximation: base + scaled length
                if max_length <= 5:
                    width = 5
                elif max_length <= 10:
                    width = 10
                elif max_length <= 20:
                    width = 15
                elif max_length <= 30:
                    width = 20
                elif max_length <= 35:
                    width = 40
                elif max_length <= 47:
                    width = 13
                else:
                    width = min(50, max_length * 0.9)  # Cap at 25

                ws.column_dimensions[column].width = width

            wb.save(output_file1)
            print(f"Auto-filters, wrapped headers, and column widths adjusted in '{sheet_name}' sheet.")
            developer_logger.info(f"Auto-filters, wrapped headers, and column widths adjusted in '{sheet_name}' sheet.")
            user_logger.info(f"Auto-filters, wrapped headers, and column widths adjusted in '{sheet_name}' sheet.")
        else:
            print(f"Sheet '{sheet_name}' not found in the workbook.")
            developer_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")
            user_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")
    Denied_Invoice()

    def Rejected_Invoice():
        # ------------------------'Denied Invoice'
        today_str = datetime.today().strftime('%m.%d.%Y')
        output_filename = f'Ariba Invoice {today_str}.xlsx'
        output_file1 = os.path.join(report_dir_path1, output_filename)

        # Load the workbook
        wb = load_workbook(output_file1)
        sheet_name = 'Rejected Invoice'

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Set header row height to 30.00
            ws.row_dimensions[1].height = 30.0

            # Apply yellow fill to headers in columns A to J
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for col in range(1, 11):
                ws.cell(row=1, column=col).fill = yellow_fill

            # Format date columns to short date format
            date_columns = ['Invoice Date', 'Date Created', 'Date Submitted']
            header_row = [cell.value for cell in ws[1]]
            for col_idx, header in enumerate(header_row, start=1):
                if header in date_columns:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        # Convert string to datetime if needed
                        if isinstance(cell.value, str):
                            try:
                                cell.value = datetime.strptime(cell.value, '%m/%d/%Y %I:%M %p')
                            except ValueError:
                                continue  # Skip if not a valid datetime string
                        # Apply short date format
                        if isinstance(cell.value, datetime):
                            cell.number_format = numbers.FORMAT_DATE_XLSX14
            # Clean and format 'Invoice Amount' column
            if 'Invoice Amount' in header_row:
                amount_col_idx = header_row.index('Invoice Amount') + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=amount_col_idx)
                    if isinstance(cell.value, str):
                        # Remove 'USD' and any non-numeric characters except dot and comma
                        cleaned_value = re.sub(r'[^0-9.,-]', '', cell.value)
                        try:
                            numeric_value = float(cleaned_value.replace(',', ''))
                            cell.value = numeric_value
                            cell.number_format = '_($* #,##0.00_)'
                        except ValueError:
                            continue

            # Define new headers starting from column K (11)
            new_headers = [
               'Preparer', 'PO Title', 'PO Status'
            ]
            bold_font = Font(bold=True)

            # Apply light green fill to new headers
            light_green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            highlight_green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')

            for i, header in enumerate(new_headers, start=11):
                cell = ws.cell(row=1, column=i, value=header)
                cell.fill = highlight_green_fill
                cell.font = bold_font

            # Save the workbook
            wb.save(output_file1)


            print(f"Updated '{sheet_name}' sheet with new columns.")
            developer_logger.info(f"Updated '{sheet_name}' sheet with new columns.")
            user_logger.info(f"Updated '{sheet_name}' sheet with new columns.")


            report_dir_path2 = os.path.join(cwd, 'Output_File', 'Data_Files')
            purchase_orderfile = f'Purchase Order {today_str}.xlsx'
            purchase_order_path = os.path.join(report_dir_path, purchase_orderfile)

            # purchase_order_path = os.path.join(report_dir_path2, purchase_orders.xlsx)
            purchase_wb = load_workbook(purchase_order_path)
            purchase_ws = purchase_wb.active

            def preparer():
                # Find columns in purchase order sheet
                start_col1 = end_col1 = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Order ID':
                        start_col1 = cell.column
                    elif cell.value == 'Preparer':
                        end_col1 = cell.column
                    if start_col1 and end_col1:
                        break

                start_col_letter1 = purchase_ws.cell(row=1, column=start_col1).column_letter
                end_col_letter1 = purchase_ws.cell(row=1, column=end_col1).column_letter
                reference_range1 = f"'{os.path.basename(purchase_order_path)}'!${start_col_letter1}:${end_col_letter1}"
                col_index_num1 = end_col1 - start_col1 + 1

                # Find columns in invoice sheet
                Ini_Match_col = Preparer_col = Supplier_col = None
                for cell in ws[1]:
                    if cell.value == 'Initially Matched Order':
                        Ini_Match_col = cell.column
                    elif cell.value == 'Preparer':
                        Preparer_col = cell.column
                    elif cell.value == 'Supplier':
                        Supplier_col = cell.column

                # Apply VLOOKUP formula to Preparer column
                if Ini_Match_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        lookup_cell = ws.cell(row=row, column=Ini_Match_col).coordinate
                        formula = f'=VLOOKUP({lookup_cell}, {reference_range1}, {col_index_num1}, 0)'
                        ws.cell(row=row, column=Preparer_col, value=formula)

                print("✅ VLOOKUP applied to 'Preparer' column.")
                developer_logger.info("✅ VLOOKUP applied to 'Preparer' column.")
                user_logger.info("✅ VLOOKUP applied to 'Preparer' column.")

                # Build fallback dictionary from purchase order sheet
                name_col = preparer_col_master = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Supplier Name':
                        name_col = cell.column
                    elif cell.value == 'Preparer':
                        preparer_col_master = cell.column
                    if name_col and preparer_col_master:
                        break

                vendor_preparer = {}
                for row in range(2, purchase_ws.max_row + 1):
                    name = purchase_ws.cell(row=row, column=name_col).value
                    preparer = purchase_ws.cell(row=row, column=preparer_col_master).value
                    if name:
                        key = ' '.join(str(name).strip().split()[:2]).lower()
                        vendor_preparer[key] = preparer

                # Apply fallback if VLOOKUP fails
                if Supplier_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        preparer_cell = ws.cell(row=row, column=Preparer_col)
                        if preparer_cell.value == '#N/A' or preparer_cell.value is None:
                            supplier_name = ws.cell(row=row, column=Supplier_col).value
                            if supplier_name:
                                supplier_key = ' '.join(str(supplier_name).strip().split()[:2]).lower()
                                matched_preparer = vendor_preparer.get(supplier_key)
                                if matched_preparer:
                                    preparer_cell.value = matched_preparer
                                else:
                                    preparer_cell.value = "Not Found"

                # Save workbook
                wb.save(output_file1)
                print("✅ Preparer column updated with fallback logic.")
                developer_logger.info("✅ Preparer column updated with fallback logic.")
                user_logger.info("✅ Preparer column updated with fallback logic.")

            preparer()

            # ------------------------col-R

            def PO_title():
                # Find columns in purchase order sheet
                start_col1 = end_col1 = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Order ID':
                        start_col1 = cell.column
                    elif cell.value == 'Title':
                        end_col1 = cell.column
                    if start_col1 and end_col1:
                        break

                start_col_letter1 = purchase_ws.cell(row=1, column=start_col1).column_letter
                end_col_letter1 = purchase_ws.cell(row=1, column=end_col1).column_letter
                reference_range1 = f"'{os.path.basename(purchase_order_path)}'!${start_col_letter1}:${end_col_letter1}"
                col_index_num1 = end_col1 - start_col1 + 1

                # Find columns in invoice sheet
                Ini_Match_col = Preparer_col = Supplier_col = None
                for cell in ws[1]:
                    if cell.value == 'Initially Matched Order':
                        Ini_Match_col = cell.column
                    elif cell.value == 'PO Title':
                        Preparer_col = cell.column
                    elif cell.value == 'Supplier':
                        Supplier_col = cell.column

                # Apply VLOOKUP formula to Preparer column
                if Ini_Match_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        lookup_cell = ws.cell(row=row, column=Ini_Match_col).coordinate
                        formula = f'=VLOOKUP({lookup_cell}, {reference_range1}, {col_index_num1}, 0)'
                        ws.cell(row=row, column=Preparer_col, value=formula)

                print("✅ VLOOKUP applied to 'PO Title' column.")
                developer_logger.info("✅ VLOOKUP applied to 'PO Title' column.")
                user_logger.info("✅ VLOOKUP applied to 'PO Title' column.")

                # Build fallback dictionary from purchase order sheet
                name_col = preparer_col_master = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Supplier Name':
                        name_col = cell.column
                    elif cell.value == 'Title':
                        preparer_col_master = cell.column
                    if name_col and preparer_col_master:
                        break

                vendor_preparer = {}
                for row in range(2, purchase_ws.max_row + 1):
                    name = purchase_ws.cell(row=row, column=name_col).value
                    preparer = purchase_ws.cell(row=row, column=preparer_col_master).value
                    if name:
                        key = ' '.join(str(name).strip().split()[:2]).lower()
                        vendor_preparer[key] = preparer

                # Apply fallback if VLOOKUP fails
                if Supplier_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        preparer_cell = ws.cell(row=row, column=Preparer_col)
                        if preparer_cell.value == '#N/A' or preparer_cell.value is None:
                            supplier_name = ws.cell(row=row, column=Supplier_col).value
                            if supplier_name:
                                supplier_key = ' '.join(str(supplier_name).strip().split()[:2]).lower()
                                matched_preparer = vendor_preparer.get(supplier_key)
                                if matched_preparer:
                                    preparer_cell.value = matched_preparer
                                else:
                                    preparer_cell.value = "Not Found"

                # Save workbook
                wb.save(output_file1)
                print("✅ PO Title column updated with fallback logic.")
                developer_logger.info("✅ PO Title column updated with fallback logic.")
                user_logger.info("✅ PO Title column updated with fallback logic.")

            PO_title()

            # ------------------------col-S

            def PO_Status():
                # Find columns in purchase order sheet
                start_col1 = end_col1 = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Order ID':
                        start_col1 = cell.column
                    elif cell.value == 'Status':
                        end_col1 = cell.column
                    if start_col1 and end_col1:
                        break

                start_col_letter1 = purchase_ws.cell(row=1, column=start_col1).column_letter
                end_col_letter1 = purchase_ws.cell(row=1, column=end_col1).column_letter
                reference_range1 = f"'{os.path.basename(purchase_order_path)}'!${start_col_letter1}:${end_col_letter1}"
                col_index_num1 = end_col1 - start_col1 + 1

                # Find columns in invoice sheet
                Ini_Match_col = Preparer_col = Supplier_col = None
                for cell in ws[1]:
                    if cell.value == 'Initially Matched Order':
                        Ini_Match_col = cell.column
                    elif cell.value == 'PO Status':
                        Preparer_col = cell.column
                    elif cell.value == 'Supplier':
                        Supplier_col = cell.column

                # Apply VLOOKUP formula to Preparer column
                if Ini_Match_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        lookup_cell = ws.cell(row=row, column=Ini_Match_col).coordinate
                        formula = f'=VLOOKUP({lookup_cell}, {reference_range1}, {col_index_num1}, 0)'
                        ws.cell(row=row, column=Preparer_col, value=formula)

                print("✅ VLOOKUP applied to 'PO Status' column.")
                developer_logger.info("✅ VLOOKUP applied to 'PO Status' column.")
                user_logger.info("✅ VLOOKUP applied to 'PO Status' column.")

                # Build fallback dictionary from purchase order sheet
                name_col = preparer_col_master = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Supplier Name':
                        name_col = cell.column
                    elif cell.value == 'Status':
                        preparer_col_master = cell.column
                    if name_col and preparer_col_master:
                        break

                vendor_preparer = {}
                for row in range(2, purchase_ws.max_row + 1):
                    name = purchase_ws.cell(row=row, column=name_col).value
                    preparer = purchase_ws.cell(row=row, column=preparer_col_master).value
                    if name:
                        key = ' '.join(str(name).strip().split()[:2]).lower()
                        vendor_preparer[key] = preparer

                # Apply fallback if VLOOKUP fails
                if Supplier_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        preparer_cell = ws.cell(row=row, column=Preparer_col)
                        if preparer_cell.value == '#N/A' or preparer_cell.value is None:
                            supplier_name = ws.cell(row=row, column=Supplier_col).value
                            if supplier_name:
                                supplier_key = ' '.join(str(supplier_name).strip().split()[:2]).lower()
                                matched_preparer = vendor_preparer.get(supplier_key)
                                if matched_preparer:
                                    preparer_cell.value = matched_preparer
                                else:
                                    preparer_cell.value = "Not Found"

                # Save workbook
                wb.save(output_file1)
                print("✅ PO Status column updated with fallback logic.")
                developer_logger.info("✅ PO Status column updated with fallback logic.")
                user_logger.info("✅ PO Status column updated with fallback logic.")

            PO_Status()
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Determine the last column with headers
                max_col = ws.max_column
                last_col_letter = ws.cell(row=1, column=max_col).column_letter

                # Apply auto-filter to the header row
                ws.auto_filter.ref = f"A1:{last_col_letter}1"

                # Save the workbook
                wb.save(output_file1)
                print(f"Auto-filters applied to all headers in '{sheet_name}' sheet.")
                developer_logger.info(f"Auto-filters applied to all headers in '{sheet_name}' sheet.")
                user_logger.info(f"Auto-filters applied to all headers in '{sheet_name}' sheet.")
            else:
                print(f"Sheet '{sheet_name}' not found in the workbook.")
                developer_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")
                user_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")
        else:
            print(f"Sheet '{sheet_name}' not found in {output_filename}.")
            developer_logger.info(f"Sheet '{sheet_name}' not found in {output_filename}.")
            user_logger.info(f"Sheet '{sheet_name}' not found in {output_filename}.")
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Apply auto-filter
            max_col = ws.max_column
            last_col_letter = ws.cell(row=1, column=max_col).column_letter
            ws.auto_filter.ref = f"A1:{last_col_letter}1"

            # Wrap text in header cells
            for col in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col)
                cell.alignment = Alignment(wrap_text=True)

            # Adjust column widths based on content
            for col in ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=max_col):
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        length = len(str(cell.value))
                        max_length = max(max_length, length)

                # Non-linear approximation: base + scaled length
                if max_length <= 5:
                    width = 5
                elif max_length <= 10:
                    width = 10
                elif max_length <= 20:
                    width = 15
                elif max_length <= 30:
                    width = 20
                elif max_length <= 35:
                    width = 40
                elif max_length <= 47:
                    width = 13
                else:
                    width = min(50, max_length * 0.9)  # Cap at 25

                ws.column_dimensions[column].width = width

            wb.save(output_file1)
            print(f"Auto-filters, wrapped headers, and column widths adjusted in '{sheet_name}' sheet.")
            developer_logger.info(f"Auto-filters, wrapped headers, and column widths adjusted in '{sheet_name}' sheet.")
            user_logger.info(f"Auto-filters, wrapped headers, and column widths adjusted in '{sheet_name}' sheet.")
        else:
            print(f"Sheet '{sheet_name}' not found in the workbook.")
            developer_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")
            user_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")
    Rejected_Invoice()

    def Reconciling_Invoice():
        # ------------------------'Denied Invoice'
        today_str = datetime.today().strftime('%m.%d.%Y')
        output_filename = f'Ariba Invoice {today_str}.xlsx'
        output_file1 = os.path.join(report_dir_path1, output_filename)

        # Load the workbook
        wb = load_workbook(output_file1)
        sheet_name = 'Reconciling Invoice'

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Set header row height to 30.00
            ws.row_dimensions[1].height = 30.0

            # Apply yellow fill to headers in columns A to J
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for col in range(1, 11):
                ws.cell(row=1, column=col).fill = yellow_fill

            # Format date columns to short date format
            date_columns = ['Invoice Date', 'Date Created', 'Date Submitted']
            header_row = [cell.value for cell in ws[1]]
            for col_idx, header in enumerate(header_row, start=1):
                if header in date_columns:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        # Convert string to datetime if needed
                        if isinstance(cell.value, str):
                            try:
                                cell.value = datetime.strptime(cell.value, '%m/%d/%Y %I:%M %p')
                            except ValueError:
                                continue  # Skip if not a valid datetime string
                        # Apply short date format
                        if isinstance(cell.value, datetime):
                            cell.number_format = numbers.FORMAT_DATE_XLSX14
            # Clean and format 'Invoice Amount' column
            if 'Invoice Amount' in header_row:
                amount_col_idx = header_row.index('Invoice Amount') + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=amount_col_idx)
                    if isinstance(cell.value, str):
                        # Remove 'USD' and any non-numeric characters except dot and comma
                        cleaned_value = re.sub(r'[^0-9.,-]', '', cell.value)
                        try:
                            numeric_value = float(cleaned_value.replace(',', ''))
                            cell.value = numeric_value
                            cell.number_format = '_($* #,##0.00_)'
                        except ValueError:
                            continue

            # Define new headers starting from column K (11)
            new_headers = [
               'Preparer', 'PO Title', 'PO Status'
            ]
            bold_font = Font(bold=True)

            # Apply light green fill to new headers
            light_green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            highlight_green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')

            for i, header in enumerate(new_headers, start=11):
                cell = ws.cell(row=1, column=i, value=header)
                cell.fill = highlight_green_fill
                cell.font = bold_font

            # Save the workbook
            wb.save(output_file1)

            print(f"Updated '{sheet_name}' sheet with new columns.")
            developer_logger.info(f"Updated '{sheet_name}' sheet with new columns.")
            user_logger.info(f"Updated '{sheet_name}' sheet with new columns.")

            report_dir_path2 = os.path.join(cwd, 'Output_File', 'Data_Files')
            purchase_orderfile = f'Purchase Order {today_str}.xlsx'
            purchase_order_path = os.path.join(report_dir_path, purchase_orderfile)

            # purchase_order_path = os.path.join(report_dir_path2, 'purchase_orders.xlsx')
            purchase_wb = load_workbook(purchase_order_path)
            purchase_ws = purchase_wb.active

            def preparer():
                # Find columns in purchase order sheet
                start_col1 = end_col1 = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Order ID':
                        start_col1 = cell.column
                    elif cell.value == 'Preparer':
                        end_col1 = cell.column
                    if start_col1 and end_col1:
                        break

                start_col_letter1 = purchase_ws.cell(row=1, column=start_col1).column_letter
                end_col_letter1 = purchase_ws.cell(row=1, column=end_col1).column_letter
                reference_range1 = f"'{os.path.basename(purchase_order_path)}'!${start_col_letter1}:${end_col_letter1}"
                col_index_num1 = end_col1 - start_col1 + 1

                # Find columns in invoice sheet
                Ini_Match_col = Preparer_col = Supplier_col = None
                for cell in ws[1]:
                    if cell.value == 'Initially Matched Order':
                        Ini_Match_col = cell.column
                    elif cell.value == 'Preparer':
                        Preparer_col = cell.column
                    elif cell.value == 'Supplier':
                        Supplier_col = cell.column

                # Apply VLOOKUP formula to Preparer column
                if Ini_Match_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        lookup_cell = ws.cell(row=row, column=Ini_Match_col).coordinate
                        formula = f'=VLOOKUP({lookup_cell}, {reference_range1}, {col_index_num1}, 0)'
                        ws.cell(row=row, column=Preparer_col, value=formula)

                print("✅ VLOOKUP applied to 'Preparer' column.")
                developer_logger.info("✅ VLOOKUP applied to 'Preparer' column.")
                user_logger.info("✅ VLOOKUP applied to 'Preparer' column.")

                # Build fallback dictionary from purchase order sheet
                name_col = preparer_col_master = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Supplier Name':
                        name_col = cell.column
                    elif cell.value == 'Preparer':
                        preparer_col_master = cell.column
                    if name_col and preparer_col_master:
                        break

                vendor_preparer = {}
                for row in range(2, purchase_ws.max_row + 1):
                    name = purchase_ws.cell(row=row, column=name_col).value
                    preparer = purchase_ws.cell(row=row, column=preparer_col_master).value
                    if name:
                        key = ' '.join(str(name).strip().split()[:2]).lower()
                        vendor_preparer[key] = preparer

                # Apply fallback if VLOOKUP fails
                if Supplier_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        preparer_cell = ws.cell(row=row, column=Preparer_col)
                        if preparer_cell.value == '#N/A' or preparer_cell.value is None:
                            supplier_name = ws.cell(row=row, column=Supplier_col).value
                            if supplier_name:
                                supplier_key = ' '.join(str(supplier_name).strip().split()[:2]).lower()
                                matched_preparer = vendor_preparer.get(supplier_key)
                                if matched_preparer:
                                    preparer_cell.value = matched_preparer
                                else:
                                    preparer_cell.value = "Not Found"

                # Save workbook
                wb.save(output_file1)
                print("✅ Preparer column updated with fallback logic.")
                developer_logger.info("✅ Preparer column updated with fallback logic.")
                user_logger.info("✅ Preparer column updated with fallback logic.")

            preparer()

            # ------------------------col-R

            def PO_title():
                # Find columns in purchase order sheet
                start_col1 = end_col1 = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Order ID':
                        start_col1 = cell.column
                    elif cell.value == 'Title':
                        end_col1 = cell.column
                    if start_col1 and end_col1:
                        break

                start_col_letter1 = purchase_ws.cell(row=1, column=start_col1).column_letter
                end_col_letter1 = purchase_ws.cell(row=1, column=end_col1).column_letter
                reference_range1 = f"'{os.path.basename(purchase_order_path)}'!${start_col_letter1}:${end_col_letter1}"
                col_index_num1 = end_col1 - start_col1 + 1

                # Find columns in invoice sheet
                Ini_Match_col = Preparer_col = Supplier_col = None
                for cell in ws[1]:
                    if cell.value == 'Initially Matched Order':
                        Ini_Match_col = cell.column
                    elif cell.value == 'PO Title':
                        Preparer_col = cell.column
                    elif cell.value == 'Supplier':
                        Supplier_col = cell.column

                # Apply VLOOKUP formula to Preparer column
                if Ini_Match_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        lookup_cell = ws.cell(row=row, column=Ini_Match_col).coordinate
                        formula = f'=VLOOKUP({lookup_cell}, {reference_range1}, {col_index_num1}, 0)'
                        ws.cell(row=row, column=Preparer_col, value=formula)

                print("✅ VLOOKUP applied to 'PO Title' column.")
                developer_logger.info("✅ VLOOKUP applied to 'PO Title' column.")
                user_logger.info("✅ VLOOKUP applied to 'PO Title' column.")

                # Build fallback dictionary from purchase order sheet
                name_col = preparer_col_master = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Supplier Name':
                        name_col = cell.column
                    elif cell.value == 'Title':
                        preparer_col_master = cell.column
                    if name_col and preparer_col_master:
                        break

                vendor_preparer = {}
                for row in range(2, purchase_ws.max_row + 1):
                    name = purchase_ws.cell(row=row, column=name_col).value
                    preparer = purchase_ws.cell(row=row, column=preparer_col_master).value
                    if name:
                        key = ' '.join(str(name).strip().split()[:2]).lower()
                        vendor_preparer[key] = preparer

                # Apply fallback if VLOOKUP fails
                if Supplier_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        preparer_cell = ws.cell(row=row, column=Preparer_col)
                        if preparer_cell.value == '#N/A' or preparer_cell.value is None:
                            supplier_name = ws.cell(row=row, column=Supplier_col).value
                            if supplier_name:
                                supplier_key = ' '.join(str(supplier_name).strip().split()[:2]).lower()
                                matched_preparer = vendor_preparer.get(supplier_key)
                                if matched_preparer:
                                    preparer_cell.value = matched_preparer
                                else:
                                    preparer_cell.value = "Not Found"

                # Save workbook
                wb.save(output_file1)
                print("✅ PO Title column updated with fallback logic.")
                developer_logger.info("✅ PO Title column updated with fallback logic.")
                user_logger.info("✅ PO Title column updated with fallback logic.")

            PO_title()

            # ------------------------col-S

            def PO_Status():
                # Find columns in purchase order sheet
                start_col1 = end_col1 = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Order ID':
                        start_col1 = cell.column
                    elif cell.value == 'Status':
                        end_col1 = cell.column
                    if start_col1 and end_col1:
                        break

                start_col_letter1 = purchase_ws.cell(row=1, column=start_col1).column_letter
                end_col_letter1 = purchase_ws.cell(row=1, column=end_col1).column_letter
                reference_range1 = f"'{os.path.basename(purchase_order_path)}'!${start_col_letter1}:${end_col_letter1}"
                col_index_num1 = end_col1 - start_col1 + 1

                # Find columns in invoice sheet
                Ini_Match_col = Preparer_col = Supplier_col = None
                for cell in ws[1]:
                    if cell.value == 'Initially Matched Order':
                        Ini_Match_col = cell.column
                    elif cell.value == 'PO Status':
                        Preparer_col = cell.column
                    elif cell.value == 'Supplier':
                        Supplier_col = cell.column

                # Apply VLOOKUP formula to Preparer column
                if Ini_Match_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        lookup_cell = ws.cell(row=row, column=Ini_Match_col).coordinate
                        formula = f'=VLOOKUP({lookup_cell}, {reference_range1}, {col_index_num1}, 0)'
                        ws.cell(row=row, column=Preparer_col, value=formula)

                print("✅ VLOOKUP applied to 'PO Status' column.")
                developer_logger.info("✅ VLOOKUP applied to 'PO Status' column.")
                user_logger.info("✅ VLOOKUP applied to 'PO Status' column.")

                # Build fallback dictionary from purchase order sheet
                name_col = preparer_col_master = None
                for cell in purchase_ws[1]:
                    if cell.value == 'Supplier Name':
                        name_col = cell.column
                    elif cell.value == 'Status':
                        preparer_col_master = cell.column
                    if name_col and preparer_col_master:
                        break

                vendor_preparer = {}
                for row in range(2, purchase_ws.max_row + 1):
                    name = purchase_ws.cell(row=row, column=name_col).value
                    preparer = purchase_ws.cell(row=row, column=preparer_col_master).value
                    if name:
                        key = ' '.join(str(name).strip().split()[:2]).lower()
                        vendor_preparer[key] = preparer

                # Apply fallback if VLOOKUP fails
                if Supplier_col and Preparer_col:
                    for row in range(2, ws.max_row + 1):
                        preparer_cell = ws.cell(row=row, column=Preparer_col)
                        if preparer_cell.value == '#N/A' or preparer_cell.value is None:
                            supplier_name = ws.cell(row=row, column=Supplier_col).value
                            if supplier_name:
                                supplier_key = ' '.join(str(supplier_name).strip().split()[:2]).lower()
                                matched_preparer = vendor_preparer.get(supplier_key)
                                if matched_preparer:
                                    preparer_cell.value = matched_preparer
                                else:
                                    preparer_cell.value = "Not Found"

                # Save workbook
                wb.save(output_file1)
                print("✅ PO Status column updated with fallback logic.")
                developer_logger.info("✅ PO Status column updated with fallback logic.")
                user_logger.info("✅ PO Status column updated with fallback logic.")

            PO_Status()
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Determine the last column with headers
                max_col = ws.max_column
                last_col_letter = ws.cell(row=1, column=max_col).column_letter

                # Apply auto-filter to the header row
                ws.auto_filter.ref = f"A1:{last_col_letter}1"

                # Save the workbook
                wb.save(output_file1)
                print(f"Auto-filters applied to all headers in '{sheet_name}' sheet.")
                developer_logger.info(f"Auto-filters applied to all headers in '{sheet_name}' sheet.")
                user_logger.info(f"Auto-filters applied to all headers in '{sheet_name}' sheet.")
            else:
                print(f"Sheet '{sheet_name}' not found in the workbook.")
                developer_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")
                user_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")
        else:
            print(f"Sheet '{sheet_name}' not found in {output_filename}.")
            developer_logger.info(f"Sheet '{sheet_name}' not found in {output_filename}.")
            user_logger.info(f"Sheet '{sheet_name}' not found in {output_filename}.")

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Apply auto-filter
            max_col = ws.max_column
            last_col_letter = ws.cell(row=1, column=max_col).column_letter
            ws.auto_filter.ref = f"A1:{last_col_letter}1"

            # Wrap text in header cells
            for col in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col)
                cell.alignment = Alignment(wrap_text=True)

            # Adjust column widths based on content
            for col in ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=max_col):
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        length = len(str(cell.value))
                        max_length = max(max_length, length)

                # Non-linear approximation: base + scaled length
                if max_length <= 5:
                    width = 5
                elif max_length <= 10:
                    width = 10
                elif max_length <= 20:
                    width = 15
                elif max_length <= 30:
                    width = 20
                elif max_length <= 35:
                    width = 40
                elif max_length <= 47:
                    width = 13
                else:
                    width = min(50, max_length * 0.9)  # Cap at 25

                ws.column_dimensions[column].width = width

            wb.save(output_file1)
            print(f"Auto-filters, wrapped headers, and column widths adjusted in '{sheet_name}' sheet.")
            developer_logger.info(f"Auto-filters, wrapped headers, and column widths adjusted in '{sheet_name}' sheet.")
            user_logger.info(f"Auto-filters, wrapped headers, and column widths adjusted in '{sheet_name}' sheet.")
        else:
            print(f"Sheet '{sheet_name}' not found in the workbook.")
            developer_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")
            user_logger.info(f"Sheet '{sheet_name}' not found in the workbook.")
    Reconciling_Invoice()

    email_file_path = os.path.join(input_dir, 'email_id.txt')
    if os.path.exists(email_file_path):
        with open(email_file_path, 'r') as f:
            email_address = f.readline().strip()

        SENDER = email_address
        email = email_address
        sendMail.send_email_consolidated(SENDER, email)

    print("Automation tasks completed successfully.")
    developer_logger.info(f"Automation tasks completed successfully.")
    user_logger.info(f"Automation tasks completed successfully.")
